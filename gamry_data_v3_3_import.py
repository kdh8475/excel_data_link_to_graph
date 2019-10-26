from openpyxl import load_workbook
#import pandas as pd
import os
#import gamry_parser as parser
import random

path_data = "./files/191016/"
path_format = r"./files/PCFC template (2019)_DHK_v1_.xlsx"
path_file1 = "./files/111.xlsx"
path_file2 = r"./files/111.xlsx"

file_list = os.listdir(path_data)
print(file_list)

book = load_workbook(path_format)
book.save(path_file1)
# book = load_workbook(path_file2)
# ws_iv = book['j-V-P']
#
#
# for files in file_list:
#     writer = pd.ExcelWriter(path_file2, engine='openpyxl')
#     writer.book = book
#
#     row = 2
#     row_iv = 5
#     row_eis = 6
#
#     if files.find('650_4_LSV') is not -1:
#         v_col = 9
#         i_col = 10
#
#         print(files)
#         file_name = path_data + files
#         ca = parser.ChronoAmperometry(to_timestamp=True)
#         ca.load(filename=file_name)
#         #print(ca.get_curve_data())
#         dta = ca.get_curve_data()
#         dta.to_excel(writer, '1', index=False)
#         work_sheet = book['1']
#
#         for tr in work_sheet:
#             v = work_sheet.cell(row=row, column=2).value
#             i = work_sheet.cell(row=row, column=3).value
#             if not v == None:
#                 ws_iv.cell(row=row_iv, column=v_col, value=v)
#                 ws_iv.cell(row=row_iv, column=i_col, value=i)
#                 row += 1
#                 row_iv += 1
#         book.remove(book['1'])
#
#     if files.find('600_LSV') is not -1:
#         v_col = 16
#         i_col = 17
#
#         print(files)
#         file_name = path_data + files
#         ca = parser.ChronoAmperometry(to_timestamp=True)
#         ca.load(filename=file_name)
#         # print(ca.get_curve_data())
#         dta = ca.get_curve_data()
#         dta.to_excel(writer, '1', index=False)
#         work_sheet = book['1']
#
#         for tr in work_sheet:
#             v = work_sheet.cell(row=row, column=2).value
#             i = work_sheet.cell(row=row, column=3).value
#             if not v == None:
#                 ws_iv.cell(row=row_iv, column=v_col, value=v)
#                 ws_iv.cell(row=row_iv, column=i_col, value=i)
#                 row += 1
#                 row_iv += 1
#         book.remove(book['1'])
#
#     if files.find('550_LSV') is not -1:
#         v_col = 23
#         i_col = 24
#
#         print(files)
#         file_name = path_data + files
#         ca = parser.ChronoAmperometry(to_timestamp=True)
#         ca.load(filename=file_name)
#         # print(ca.get_curve_data())
#         dta = ca.get_curve_data()
#         dta.to_excel(writer, '1', index=False)
#         work_sheet = book['1']
#
#         for tr in work_sheet:
#             v = work_sheet.cell(row=row, column=2).value
#             i = work_sheet.cell(row=row, column=3).value
#             if not v == None:
#                 ws_iv.cell(row=row_iv, column=v_col, value=v)
#                 ws_iv.cell(row=row_iv, column=i_col, value=i)
#                 row += 1
#                 row_iv += 1
#         book.remove(book['1'])
#
#     if files.find('500_2_LSV') is not -1:
#         v_col = 30
#         i_col = 31
#
#         print(files)
#         file_name = path_data + files
#         ca = parser.ChronoAmperometry(to_timestamp=True)
#         ca.load(filename=file_name)
#         # print(ca.get_curve_data())
#         dta = ca.get_curve_data()
#         dta.to_excel(writer, '1', index=False)
#         work_sheet = book['1']
#
#         for tr in work_sheet:
#             v = work_sheet.cell(row=row, column=2).value
#             i = work_sheet.cell(row=row, column=3).value
#             if not v == None:
#                 ws_iv.cell(row=row_iv, column=v_col, value=v)
#                 ws_iv.cell(row=row_iv, column=i_col, value=i)
#                 row += 1
#                 row_iv += 1
#         book.remove(book['1'])
#
# writer.save()
# writer.close()